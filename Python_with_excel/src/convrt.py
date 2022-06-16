from openpyxl import Workbook, load_workbook
import timeit


def main():
    start = timeit.default_timer()
    
    sheet = load_workbook('excel_files/experimental.xlsx')
    
    stop = timeit.default_timer()
    total = stop - start
    
    print(f"It took {total} seconds to open the workbook")
    
    
    pages_remove = []
    sheet_name1 = sheet.sheetnames
    for i in range(1, len(sheet_name1) + 1):
        active_sheet = sheet[f'Table {i}']
        if active_sheet['A1'].value == "Shop Information" or "CASH MEMO" in active_sheet['A1'].value:
            to_remove = sheet[f'Table {i}']
            pages_remove.append(i)
            sheet.remove(to_remove)
            
    
    new_sheet_name = sheet.sheetnames
    for sheets in new_sheet_name:
        print(sheets)
        
    sheet_name2 = sheet.sheetnames
    for pages in pages_remove:
        print(pages)
    
    
main()