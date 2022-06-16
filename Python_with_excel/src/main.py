from openpyxl import load_workbook
import timeit


def from_pdf_excel():
    # ? This is loading the converted pdf file into the program!

    print("Opening all the excel files... This might take few seconds...")
    # * Invoice
    converted = load_workbook('excel_files/invoice.xlsx')
    invoice = converted.active

    # * Copying CNIC
    cnic_Sheet = load_workbook('excel_files/FBR.xlsx')
    cnic = cnic_Sheet.active

    # * New Excel File
    new_file = load_workbook('excel_files/sample.xlsx')
    extract = new_file.active

    #! 6 counters will be needed
    #! 1 --> Buyer CNIC
    #! 2 --> Buyer Name
    #! 3 --> Document Numbers / Invoice Numbers
    #! 4 --> Document Date / Invoice Date
    #! 5 --> Total Quantity --> if 1 X 5 --> ctn qty else pcs qty
    #! 6 --> Sales tax

    # * Counters
    invoice_sheet_inc = 1
    buyer_name_inc = 2
    document_date_inc = 2
    invoice_num_inc = 2
    sales_tax_inc = 2
    invoice_sales_inc = 3
    invoice_quantity_inc = 2
    quant_inc = 2
    cnic_inc = 2
    export_inc = 2

    shop_names = []  # ? So that the program does not have to load the files from the exported file again!
    #! Placing Names, Document Invoice Number & Document Invoice Date portion
    print("Placing Names...")
    for tables in range(1, len(converted.sheetnames) + 1):
        if tables == len(converted.sheetnames):
            continue
        if invoice_sheet_inc + 3 > len(converted.sheetnames) + 3:
            break
        else:
            # ? Extracting demo-data from table 1 currently
            invoice = converted[f'Table {invoice_sheet_inc}']
            if invoice['A1'].value == "CASH MEMO / SALE INVOICE":
                invoice_sheet_inc += 2
                continue
            # * This is the fix for dashes!
            else:
                strings = list(invoice['B1'].value)
                if '\n' in strings:
                    strings = "".join(strings[strings.index('-') + 1:strings.index('\n')])
                else:
                    strings = "".join(strings[strings.index('-') + 1::])

            # ? Assigning names from tables by using appropriate incrementing of rows
            strings = strings.replace('G/S', 'GENERAL STORE')
            strings = strings.replace('S/S', 'SUPER STORE')
            strings = strings.replace('C/C', 'CASH AND CARRY')
            strings = strings.replace('K/S', 'KARYANA STORE')
            strings = strings.replace('G.STORE', 'GENERAL STORE')
            strings = strings.replace('C&C', 'CASH AND CARRY')
            strings = strings.replace('GS', 'GENERAL STORE')
            strings = strings.replace('Cash & Carry', 'CASH AND CARRY')
            strings = strings.replace('Cash & carry', 'CASH AND CARRY')
            strings = strings.replace('cash & carry', 'CASH AND CARRY')
            

            strings = strings.upper()

            shop_names.append(strings)
            extract[f'D{buyer_name_inc}'].value = strings

            # ? Assigning document dates & invoices by using appropriate incrementing of rows
            if "Date" in str(invoice['F1'].value):
                if '\n' in str(invoice['G1'].value):
                    invoice_date = str(invoice['G1'].value)
                    print(invoice_date, end = ' ')
                    invoice_date = list(invoice_date[:invoice_date.index('\n')])
                    invoice_date = "".join(invoice_date)

                    invoice_number = str(invoice['G1'].value)
                    print(invoice_number, tables)
                    invoice_number = list(invoice_number[invoice_number.index('-') + 1:])
                    invoice_number = "".join(invoice_number)
                else:
                    invoice_date = str(invoice['G1'].value)
                    invoice_date = invoice_date[:invoice_date.index('00:')]
                    invoice_date = invoice_date.replace('-', '/')

                    invoice_number = str(invoice['G2'].value)
                    invoice_number = list(invoice_number[invoice_number.index('-') + 1:])
                    invoice_number = "".join(invoice_number)
            elif '\n' in str(invoice['F1'].value)  and not "Date" in str(invoice['F1'].value):
                invoice_date = str(invoice['F1'].value)
                print(invoice_date, end = ' ')
                invoice_date = list(invoice_date[:invoice_date.index('\n')])
                invoice_date = "".join(invoice_date)

                invoice_number = str(invoice['F1'].value)
                print(invoice_number, tables)
                invoice_number = list(invoice_number[invoice_number.index('-') + 1:])
                invoice_number = "".join(invoice_number)
            else:
                invoice_date = str(invoice['F1'].value)
                invoice_date = invoice_date[:invoice_date.index('00:')]
                invoice_date = invoice_date.replace('-', '/')

                invoice_number = str(invoice['F2'].value)
                invoice_number = list(invoice_number[invoice_number.index('-') + 1:])
                invoice_number = "".join(invoice_number)

            # ? Assigning invoice date and invoice number from tables by using appropriate incrementing of rows
            extract[f'I{document_date_inc}'].value = invoice_date
            extract[f'H{invoice_num_inc}'].value = int(invoice_number)

            # ? 3 out of 6 incrementor
            buyer_name_inc += 1
            document_date_inc += 1
            invoice_num_inc += 1

            # ? Table incrementor
            invoice_sheet_inc += 3

    #! Sales Tax Portion
    print("Sales Tax portion")
    for sale in range(1, len(converted.sheetnames) + 1):
        while invoice_sales_inc + 3 <= len(converted.sheetnames) + 1:
            invoice = converted[f'Table {invoice_sales_inc}']
            if invoice['A1'].value != None:
                invoice_sales_inc += 2
                continue
            invoice_sales_inc += 3
            sales = invoice['D6'].value

            extract[f'O{sales_tax_inc}'].value = int(sales)
            sales_tax_inc += 1

    #! Quantity Portion
    print("Quantity Portion")
    for j in range(1, len(converted.sheetnames) + 1):
        while invoice_quantity_inc + 3 <= len(converted.sheetnames) + 3:
            invoice = converted[f'Table {invoice_quantity_inc}']
            if invoice['B1'].value == "Distributor Information                                                                                 Invoice":
                invoice_quantity_inc += 2
                continue
            Quantity = 0
            # ? This iterates for the product names
            for i in range(2, len(invoice['B'])):
                product_name = invoice[f'B{i}'].value
                if "X 5" in product_name or "x 5" in product_name or "x5" in product_name or "X5" in product_name:
                    Quantity += invoice[f'E{i}'].value
                else:
                    Quantity += invoice[f'F{i}'].value
                    

            extract[f'M{quant_inc}'].value = Quantity
            quant_inc += 1
            invoice_quantity_inc += 3

    #! CNIC Portion
    print("Looking for cnics")
    CNIC_List = []
    while cnic[f'D{cnic_inc}'].value != None:
        cnic_name = cnic[f'D{cnic_inc}'].value
        cnic_number = cnic[f'C{cnic_inc}'].value
        CNIC_List.append([cnic_number, cnic_name])
        cnic_inc += 1

    # ? Loading names from exported file!
    for name in shop_names:
        name = name.replace('G/S', 'GENERAL STORE')
        name = name.replace('S/S', 'SUPER STORE')
        name = name.replace('C/C', 'CASH AND CARRY')
        name = name.replace('K/S', 'KARYANA STORE')
        name = name.replace('G.STORE', 'GENERAL STORE')
        name = name.replace('C&C', 'CASH AND CARRY')
        name = name.replace('GS', 'GENERAL STORE')

        name = name.upper()

        for cnics in CNIC_List:
            if name in cnics:
                extract[f'C{export_inc}'].value = cnics[0]
                break
        export_inc += 1

    print("Saving...")
    save_start = timeit.default_timer()
    new_file.save('excel_files/Exported.xlsx')
    stop_start = timeit.default_timer()

    total_save_time = stop_start - save_start
    print(f"It took {total_save_time} seconds to save...")


def main():
    # ? Gets the start point
    start = timeit.default_timer()

    #! Actual Code
    from_pdf_excel()

    # ? Gets the end point
    stop = timeit.default_timer()

    # ? Extracts the time
    execution_time = stop - start
    # It returns time in seconds
    print(f"Program Executed in {execution_time} secs...")
    input("Press any key to close...")


main()