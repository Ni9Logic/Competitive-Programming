from linecache import cache
from openpyxl import load_workbook
from timeit import default_timer

class invoicee:
    def __init__(self):
        self.buyer_name = ''
        self.buyer_cnic = ''
        self.invoice_date = ''
        self.invoice_number = 0
        self.total_quantity = 0
        self.ValueAfterTax = 0
        
        
def insert_dash(string, index):
    return string[:index] + '-' + string[index:]

def program():
    Invoice_objects = []
    
    print("Opening Excel files this may take few seconds...")
    # * Invoice
    converted = load_workbook('excel_files/invoice.xlsx')
    invoice_sheet = converted.active
    
    # * File to export
    export_sheet = load_workbook('excel_files/sample.xlsx')
    to_export_sheet = export_sheet.active
    
    sheet_divider = 1
    #? Creating Objects...
    print("Creating Objects...")
    invoice = invoicee()
    for tables in range(1, len(converted.sheetnames) + 1):
        invoice_sheet = converted[f'Table {tables}']
        cell_value_A1 = invoice_sheet['A1'].value
        
        if cell_value_A1 == None: #! --> Sales Tax
            invoice.ValueAfterTax = invoice_sheet['D6'].value
        elif 'Name' in cell_value_A1: #! --> Name, cnic, invoice_date, invoice_number
                buyer_name = list(invoice_sheet['B1'].value)
                if '\n' in buyer_name:
                    buyer_name = "".join(buyer_name[buyer_name.index('-') + 1:buyer_name.index('\n')])
                else:
                    buyer_name = "".join(buyer_name[buyer_name.index('-') + 1::])

                # ? Assigning names from tables by using appropriate incrementing of rows
                buyer_name = buyer_name.upper()
                invoice.buyer_name = buyer_name
                invoice.buyer_name = invoice.buyer_name.replace('G/S', 'GENERAL STORE')
                invoice.buyer_name = invoice.buyer_name.replace('S/S', 'SUPER STORE')
                invoice.buyer_name = invoice.buyer_name.replace('C/C', 'CASH AND CARRY')
                invoice.buyer_name = invoice.buyer_name.replace('K/S', 'KARYANA STORE')
                invoice.buyer_name = invoice.buyer_name.replace('G.STORE', 'GENERAL STORE')
                invoice.buyer_name = invoice.buyer_name.replace('C&C', 'CASH AND CARRY')
                invoice.buyer_name = invoice.buyer_name.replace(' GS', 'GENERAL STORE')
                invoice.buyer_name = invoice.buyer_name.replace('GS', 'GENERAL STORE')
                invoice.buyer_name = invoice.buyer_name.replace('Cash & Carry', 'CASH AND CARRY')
                invoice.buyer_name = invoice.buyer_name.replace('Cash & carry', 'CASH AND CARRY')
                invoice.buyer_name = invoice.buyer_name.replace('cash & carry', 'CASH AND CARRY')
                invoice.buyer_name = invoice.buyer_name
                
                # ? Assigning CNIC
                buyer_cnic = str(invoice_sheet['B4'].value)
                
                if '_' in buyer_cnic:
                    buyer_cnic = buyer_cnic.replace('_', '-')
                elif '-' not in buyer_cnic and '_' not in buyer_cnic:
                    buyer_cnic = insert_dash(buyer_cnic, 5)
                    buyer_cnic = insert_dash(buyer_cnic, 13)
                    
                invoice.buyer_cnic = buyer_cnic
                
        elif 'Product Code' in cell_value_A1: #! --> Quantity
            for product_i in range(2, len(invoice_sheet['B'])): #? This will iterate it through the product names.
                product_name = invoice_sheet[f'B{product_i}'].value
                if "X 5" in product_name or "x 5" in product_name or "x5" in product_name or "X5" in product_name:
                    invoice.total_quantity += invoice_sheet[f'E{product_i}'].value
                else:
                    invoice.total_quantity += invoice_sheet[f'F{product_i}'].value
        else: #! --> Sales meme invoice & shop information
            continue
        
        #* Important
        sheet_divider = (sheet_divider % 3) + 1  
        if sheet_divider == 1:
            Invoice_objects.append(invoice)
            invoice = invoicee()

    for i in range(0, len(Invoice_objects)):
        print(f"-------Invoice {i + 1}-------")
        print(f"Name of the buyer is: {Invoice_objects[i].buyer_name}")
        print(f"CNIC of the buyer is: {Invoice_objects[i].buyer_cnic}")
        print(f"Quantity of the buyer is: {Invoice_objects[i].total_quantity}")
        print(f"Sales Tax of the buyer is: {Invoice_objects[i].ValueAfterTax}")
        
        
        # export_sheet.save("excel_files/Exported_v2.xlsx")
    

def main():
    start = default_timer()
    program()
    stop = default_timer()
    
    total = stop - start
    
    print(f"Your program took total of {total} seconds to complete...")
    
    
main()