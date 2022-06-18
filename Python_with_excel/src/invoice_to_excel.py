from linecache import cache
from timeit import default_timer
from openpyxl import load_workbook

class invoicee:
    def __init__(self):
        self.buyer_name = ''
        self.buyer_cnic = ''
        self.invoice_date = ''
        self.invoice_number = 0
        self.total_quantity = 0
        self.ValueAfterTax = 0
        self.NTN = 0
        
    def correct_buyer_names(self):
            self.buyer_name = self.buyer_name.replace('G/S', 'GENERAL STORE')
            self.buyer_name = self.buyer_name.replace('S/S', 'SUPER STORE')
            self.buyer_name = self.buyer_name.replace('C/C', 'CASH AND CARRY')
            self.buyer_name = self.buyer_name.replace('K/S', 'KARYANA STORE')
            self.buyer_name = self.buyer_name.replace('G.STORE', 'GENERAL STORE')
            self.buyer_name = self.buyer_name.replace('C&C', 'CASH AND CARRY')
            self.buyer_name = self.buyer_name.replace(' GS', 'GENERAL STORE')
            self.buyer_name = self.buyer_name.replace('GS', 'GENERAL STORE')
            self.buyer_name = self.buyer_name.replace('Cash & Carry', 'CASH AND CARRY')
            self.buyer_name = self.buyer_name.replace('Cash & carry', 'CASH AND CARRY')
            self.buyer_name = self.buyer_name.replace('cash & carry', 'CASH AND CARRY')
            
            return self.buyer_name
        
        
def program():
    Invoice_objects = [] #? List of objects
    
    print("Opening Excel files this may take few seconds...")
    # * Invoice
    converted = load_workbook('excel_files/invoice.xlsx') #? Loading converted pdf's invoice.
    invoice_sheet = converted.active 
    
    sheet_divider = 1 #! This is very useful --> It basically identifies which sheet we are in currently.
    #? Creating Objects...
    print("Creating Objects...")
    invoice = invoicee()
    for tables in range(1, len(converted.sheetnames) + 1): #* 1 --> Length of all the worksheets we got after converting the pdf into excel.
        invoice_sheet = converted[f'Table {tables}'] #? To get all the data in invoice_sheet's object from current table.
        cell_value_A1 = invoice_sheet['A1'].value #* This makes it further identifiable on which sheet we are currently.
        
        if cell_value_A1 == None: #! --> Sales Tax Sheet very easy.
            invoice.ValueAfterTax = invoice_sheet['D6'].value
        elif 'Name' in cell_value_A1: #! --> Main Sheet which has name, cnic, date, invoice_number, NTN
                buyer_name = list(invoice_sheet['B1'].value)
                if '\n' in buyer_name:
                    buyer_name = "".join(buyer_name[buyer_name.index('-') + 1:buyer_name.index('\n')]) #? On new line is address we don't want that and there's a dash as well
                else:                                                                                  
                    buyer_name = "".join(buyer_name[buyer_name.index('-') + 1::])   #? We don't want the dash from the name.

                # ? Assigning names from tables by using appropriate incrementing of rows
                buyer_name = buyer_name.upper()
                invoice.buyer_name = buyer_name
                invoice.buyer_name = invoice.correct_buyer_names()
                
                # ? Assigning CNIC to object
                buyer_cnic = str(invoice_sheet['B4'].value)
                if '_' in buyer_cnic:
                    buyer_cnic = buyer_cnic.replace('_', '')
                elif '-' in buyer_cnic:
                    buyer_cnic = buyer_cnic.replace('-', '')
                elif buyer_cnic == '' or buyer_cnic == None or buyer_cnic == 'None':
                    buyer_cnic == '0'
                    
                invoice.buyer_cnic = buyer_cnic
                
                
                # ? Assigning Dates to object
                if "Date" not in str(invoice_sheet['F1'].value): #? In some sheets the 'DATE' gets into F1 & Values move onto the next cell so this is the fix.
                    if '\n' in str(invoice_sheet['F1'].value): #? '\n' because the cells usually gets merged so to differentiate whats date and whats invoice number.
                        date = str(invoice_sheet['F1'].value)
                        date = list(date[:date.index('\n')])  
                        date = "".join(date)
                        invoice.invoice_date = date
                    else:
                        date = str(invoice_sheet['F1'].value)
                        date = list(date[:date.index('00:')]) #? Dates which are directly taken from the cell include '00:00:00' with them so we ignore that.
                        date = "".join(date)
                        date = date.replace('-', '/') #? These dates have dashes in them instead of '/' so we replace the dashes with '/'
                        year = date[:4:] #? These dates format are eg. 2022/05/31 which is wrong --> year = 2022
                        day = date[8:] #? Day = 31 
                        
                        #* Swapping / Reshaping the dates
                        date = date.replace(day, '!')
                        date = date.replace(year, day)
                        date = date.replace('!', year)
                        date = date.replace(' ', '')
                        
                        invoice.invoice_date = date
                else: # ? This means data cell is shifted towards right and now we have to pick the data from Cell G.
                    if '\n' in str(invoice_sheet['G1'].value):
                        date = str(invoice_sheet['G1'].value)
                        date = list(date[:date.index('\n')])
                        date = "".join(date)
                        invoice.invoice_date = date
                    else:
                        date = str(invoice_sheet['G1'].value)
                        date = list(date[:date.index('00:')])
                        date = "".join(date)
                        date = date.replace('-', '/')
                        year = date[:4:]
                        day = date[8:]
                        date = date.replace(day, '!')
                        date = date.replace(year, day)
                        date = date.replace('!', year)
                        date = date.replace(' ', '')
                        invoice.invoice_date = date
                
                # ? Assigning Invoice Numbers
                if "Date" not in str(invoice_sheet['F1'].value): #? In some sheets the 'DATE' gets into F1 & Values move onto the next cell so this is the fix.
                    if '\n' in str(invoice_sheet['F1'].value): #? '\n' because the cells usually gets merged so to differentiate whats date and whats invoice number.
                        number = str(invoice_sheet['F1'].value)
                        number = list(number[number.index('-') + 1:]) #? Invoice numbers only include digits and 'I-' is skipped.
                        number = "".join(number)
                        invoice.invoice_number = number
                    else: #? '\n' not found it means cells are not merged and we can just pick the value from the cell directly
                        number = str(invoice_sheet['F2'].value)
                        number = list(number[number.index('-') + 1:])
                        number = "".join(number)
                        invoice.invoice_number = number
                else: # ? This means data cell is shifted towards right and now we have to pick the data from Cell G.
                    if '\n' in str(invoice_sheet['G1'].value):
                        number = str(invoice_sheet['G1'].value)
                        number = list(number[number.index('-') + 1:])
                        number = "".join(number)
                        invoice.invoice_number = number
                    else:
                        number = str(invoice_sheet['G2'].value)
                        number = list(number[number.index('-') + 1:])
                        number = "".join(number)
                        invoice.invoice_number = number
                
                #? Assigning NTN
                if '\n' in str(invoice_sheet['B2'].value):
                    NTN = str(invoice_sheet['B2'].value)
                    NTN = list(NTN[NTN.index('\n') + 1:])
                    NTN = "".join(NTN)
                    invoice.NTN = NTN
                else:
                    NTN = str(invoice_sheet['B2'].value)
                    invoice.NTN = NTN
                    
                        
        elif 'Product Code' in cell_value_A1: #! --> Quantity's Sheet
            for product_i in range(2, len(invoice_sheet['B'])): #? This will iterate it through the product names.
                product_name = invoice_sheet[f'B{product_i}'].value
                if "X 5" in product_name or "x 5" in product_name or "x5" in product_name or "X5" in product_name:
                    invoice.total_quantity += invoice_sheet[f'E{product_i}'].value
                else:
                    invoice.total_quantity += invoice_sheet[f'F{product_i}'].value
                    
                    
        else: #! --> Sales meme invoice & shop information Sheet --> Useless & Time waste
            continue
        
        #* Important
        sheet_divider = (sheet_divider % 3) + 1  #? (1 % 3) + 1 = 2, (2 % 3) + 1 = 3, (3 % 3) + 1 = 1 --> We are using '% 3' because we are reading data from three sheets.
        if sheet_divider == 1: #? (2 != 1), (3 != 1), (1 == 1)
            Invoice_objects.append(invoice)
            invoice = invoicee()



    #? Exporting the object's values into the excel file...
    # * File to export
    export_sheet = load_workbook('excel_files/sample.xlsx') #? File to Save later on
    to_export_sheet = export_sheet.active
    
    print("Generating the third excel file...")

    rows = 2
    for i in range(0, len(Invoice_objects)):
        if Invoice_objects[i].buyer_cnic == '0' or Invoice_objects[i].buyer_cnic == None:             
            continue
        to_export_sheet[f'D{rows}'].value = Invoice_objects[i].buyer_name
        to_export_sheet[f'C{rows}'].value = Invoice_objects[i].buyer_cnic
        to_export_sheet[f'H{rows}'].value = int(Invoice_objects[i].invoice_number)
        to_export_sheet[f'I{rows}'].value = Invoice_objects[i].invoice_date
        to_export_sheet[f'M{rows}'].value = Invoice_objects[i].total_quantity
        to_export_sheet[f'O{rows}'].value = Invoice_objects[i].ValueAfterTax
        rows += 1
        
    print("Saving...")    
    export_sheet.save("excel_files/Exported_v2.xlsx")
    

def main():
    start = default_timer()
    program()
    stop = default_timer()
    
    total = stop - start
    
    print(f"Your program took total of {total} seconds to complete...")
    
    
main()