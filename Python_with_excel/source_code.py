from timeit import default_timer
from openpyxl import load_workbook
from loader import Loader
import pymysql

class invoicee:
    def __init__(self):
        self.buyer_name = ''
        self.buyer_cnic = ''
        self.invoice_date = ''
        self.invoice_number = 0
        self.total_quantity = 0
        self.ValueAfterTax = 0
        self.NTN = ''   
        
    def correct_buyer_names(self):
            self.buyer_name = self.buyer_name.replace('G/S', 'GENERAL STORE')
            self.buyer_name = self.buyer_name.replace('g/s', ' GENERAL STORE')
            self.buyer_name = self.buyer_name.replace('S/S', 'SUPER STORE')
            self.buyer_name = self.buyer_name.replace('C/C', 'CASH AND CARRY')
            self.buyer_name = self.buyer_name.replace('K/S', 'KARYANA STORE')
            self.buyer_name = self.buyer_name.replace('G.STORE', 'GENERAL STORE')
            self.buyer_name = self.buyer_name.replace('C&C', 'CASH AND CARRY')
            self.buyer_name = self.buyer_name.replace(' GS', 'GENERAL STORE')
            self.buyer_name = self.buyer_name.replace('GS', ' GENERAL STORE')
            self.buyer_name = self.buyer_name.replace('gs', ' GENERAL STORE')
            self.buyer_name = self.buyer_name.replace('Cash & Carry', 'CASH AND CARRY')
            self.buyer_name = self.buyer_name.replace('Cash & carry', 'CASH AND CARRY')
            self.buyer_name = self.buyer_name.replace('cash & carry', 'CASH AND CARRY')
            self.buyer_name = self.buyer_name.replace('W/S', 'WHOLE SALE')
            self.buyer_name = self.buyer_name.replace('w/s', 'WHOLE SALE')
            self.buyer_name = self.buyer_name.replace('TRD.', 'TRADER')
            self.buyer_name = self.buyer_name.replace('trd.', 'TRADER')
            self.buyer_name = self.buyer_name.replace('TRD', 'TRADER')
            self.buyer_name = self.buyer_name.replace('trd', 'TRADER')
            return self.buyer_name
    def correct_dates(self):
        self.invoice_date = list(self.invoice_date)
        month = "".join(self.invoice_date[3:5])

        if self.invoice_date[0] == '0':
            
            del self.invoice_date[0]
            del self.invoice_date[2:4]
            del self.invoice_date[3:5]
            
            
            if month == '01':
                self.invoice_date.insert(2, 'Jan')
            elif month == '02':
                self.invoice_date.insert(2, 'Feb')
            elif month == '03':
                self.invoice_date.insert(2, 'Mar')
            elif month == '04':
                self.invoice_date.insert(2, 'Apr')
            elif month == '05':
                self.invoice_date.insert(2, 'May')
            elif month == '06':
                self.invoice_date.insert(2, 'Jun')
            elif month == '07':
                self.invoice_date.insert(2, 'Jul')
            elif month == '08':
                self.invoice_date.insert(2, 'Aug')
            elif month == '09':
                self.invoice_date.insert(2, 'Sep')
            elif month == '10':
                self.invoice_date.insert(2, 'Oct')
            elif month == '11':
                self.invoice_date.insert(2, 'Nov')
            elif month == '12':
                self.invoice_date.insert(2, 'Dec') #? 01--2022               
        else:
            del self.invoice_date[3:5] #? 01--2022
            del self.invoice_date[4:6] #? 01--22
            
            if month == '01':
                self.invoice_date.insert(3, 'Jan')
            elif month == '02':
                self.invoice_date.insert(3, 'Feb')
            elif month == '03':
                self.invoice_date.insert(3, 'Mar')
            elif month == '04':
                self.invoice_date.insert(3, 'Apr')
            elif month == '05':
                self.invoice_date.insert(3, 'May')
            elif month == '06':
                self.invoice_date.insert(3, 'Jun')
            elif month == '07':
                self.invoice_date.insert(3, 'Jul')
            elif month == '08':
                self.invoice_date.insert(3, 'Aug')
            elif month == '09':
                self.invoice_date.insert(3, 'Sep')
            elif month == '10':
                self.invoice_date.insert(3, 'Oct')
            elif month == '11':
                self.invoice_date.insert(3, 'Nov')
            elif month == '12':
                self.invoice_date.insert(3, 'Dec')

        self.invoice_date = "".join(self.invoice_date)
        return self.invoice_date
    def ctrl_f(cur, shop_name : str, shop_cnic : str, fbr_shop_names : list, fbr_shop_cnics : list) -> str:
        query = "INSERT INTO fbr (Buyer_CNIC, Buyer_Name) VALUES (%s, %s)"
    
        for fbr_shop_name, fbr_shop_cnic in zip(fbr_shop_names, fbr_shop_cnics):
            if fbr_shop_name is None or fbr_shop_cnic is None:
                continue
            if shop_name == fbr_shop_name and shop_cnic == '0' or shop_cnic == 'None':
                fbr_shop_cnic = fbr_shop_cnic.replace('-', '')
                shop_cnic = fbr_shop_cnic
            elif shop_name not in fbr_shop_names and shop_cnic != '0': #? This is just to update the FBR database with time to time.
                cur.execute(query, (shop_cnic, shop_name))
                break
            #? Import data and update from the database.
            
        return shop_cnic          

def checkInt(str):
    try:
        int(str)
        return True
    except ValueError:
        return False
def program():
    loader = Loader("Connecting to FBR database...", "Connected to Fbr database!", 0.05).start()
    # To connect MySQL database
    conn = pymysql.connect(
        host='localhost',
        user='root', 
        password = "root",
        db='fbr',
        )
    
    loader.stop()
    
    
    Invoice_objects = [] #? List of objects
    
    # * Invoice
    loader = Loader("Opening the Invoice.xlsx file...", "Opened the Invoice.xlsx file!", 0.05).start()
    converted = load_workbook('excel_files/invoice.xlsx') #? Loading converted pdf's invoice.
    invoice_sheet = converted.active 
    loader.stop()
    #* FBR DB
    loader = Loader("Loading Data from FBR database...", "Done loading data from FBR Database!", 0.05).start()
    cur = conn.cursor()
    
    query = "SELECT Buyer_Name from fbr"
    cur.execute(query)
    fbr_shop_names = [item[0] for item in cur.fetchall()]
    query = "SELECT Buyer_CNIC from fbr"
    cur.execute(query)
    fbr_shop_cnics = [item[0] for item in cur.fetchall()]
        
    loader.stop()
    sheet_divider = 1 #! This is very useful --> There are multiple work sheets in Invoice File so this helps us create a single object from three work sheets.
    
    #? Creating Objects...
    loader = Loader("Storing each invoice's data in its separate object...", "Objects Created!", 0.05).start()
    invoice = invoicee()
    
    for tables in range(1, len(converted.sheetnames) + 1): #* 1 --> Length of all the worksheets we got after converting the pdf into excel.
        invoice_sheet = converted[f'Table {tables}'] #? To get all the data in invoice_sheet's object from current table.
        cell_value_A1 = invoice_sheet['A1'].value #* This makes it further identifiable on which sheet we are currently.
        
        if 'Name' in cell_value_A1: #! --> Main Sheet which has name, cnic, date, invoice_number, NTN
                buyer_name = list(invoice_sheet['B1'].value)
                if '\n' in buyer_name:
                    buyer_name = "".join(buyer_name[buyer_name.index('-') + 1:buyer_name.index('\n')]) #? On new line is address we don't want that and there's a dash as well
                else:                                                                                  
                    buyer_name = "".join(buyer_name[buyer_name.index('-') + 1::])   #? We don't want the dash in the file.

                # ? Assigning names from tables by using appropriate incrementing of rows
                buyer_name = buyer_name.upper()
                invoice.buyer_name = buyer_name
                #? This standards the name replaces g/s with general stores etc.
                invoice.buyer_name = invoice.correct_buyer_names()
                
                # ? Assigning CNIC to object
                buyer_cnic = str(invoice_sheet['B4'].value)
                if '_' in buyer_cnic: #? Some had underscores we remove them.
                    buyer_cnic = buyer_cnic.replace('_', '')
                elif '-' in buyer_cnic:
                    buyer_cnic = buyer_cnic.replace('-', '') #? Some had dashes we also don't want that.
                elif buyer_cnic == '' or buyer_cnic == None or buyer_cnic == 'None':
                    buyer_cnic == '0'
                    
                invoice.buyer_cnic = buyer_cnic
                
                
                # ? Assigning Dates to object
                if "Date" not in str(invoice_sheet['F1'].value): #? In some sheets the 'DATE' gets into F1 & Values move onto the next cell so this is the fix.
                    if '\n' in str(invoice_sheet['F1'].value): #? '\n' because the cells usually gets merged so to differentiate whats date and whats invoice number.
                        date = str(invoice_sheet['F1'].value)
                        date = list(date[:date.index('\n')])  
                        
                        date = "".join(date)
                        date = date.replace('/', '-')
                        
                        invoice.invoice_date = date
                        invoice.invoice_date = invoice.correct_dates()
                        
                    else:
                        date = str(invoice_sheet['F1'].value)
                        date = list(date[:date.index('00:')]) #? Dates which are directly taken from the cell include '00:00:00' with them so we ignore that.
                        date = "".join(date)
                        date = date.replace('/', '-') #? These dates have dashes in them instead of '/' so we replace the dashes with '/'
                        
                        year = date[:4:] #? These dates format are eg. 2022/05/31 which is wrong --> year = 2022
                        day = date[8:] #? Day = 31 
                        
                        #* Swapping / Reshaping the dates
                        date = date.replace(day, '!')
                        date = date.replace(year, day)
                        date = date.replace('!', year)
                        date = date.replace(' ', '')
                        
                        invoice.invoice_date = date
                        invoice.invoice_date = invoice.correct_dates()
                        
                else: # ? This means data cell is shifted towards right and now we have to pick the data from Cell G.
                    if '\n' in str(invoice_sheet['G1'].value):
                        date = str(invoice_sheet['G1'].value)
                        date = list(date[:date.index('\n')])
                        
                        date = "".join(date)
                        date = date.replace('/', '-')
                        
                        invoice.invoice_date = date
                        invoice.invoice_date = invoice.correct_dates()
                    else:
                        date = str(invoice_sheet['G1'].value)
                        date = list(date[:date.index('00:')])
                        date = "".join(date)
                        date = date.replace('/', '-')
                        
                        year = date[:4:]
                        day = date[8:]
                        
                        #* Swapping / Reshaping the dates
                        date = date.replace(day, '!')
                        date = date.replace(year, day)
                        date = date.replace('!', year)
                        date = date.replace(' ', '')
                        
                        invoice.invoice_date = date
                        invoice.invoice_date = invoice.correct_dates()
                        
                
                # ? Assigning Invoice Numbers
                if "Date" not in str(invoice_sheet['F1'].value): #? In some sheets the 'DATE' gets into F1 & Values move onto the next cell so this is the fix.
                    if '\n' in str(invoice_sheet['F1'].value): #? '\n' because the cells usually gets merged so to differentiate whats date and whats invoice number.
                        number = str(invoice_sheet['F1'].value)
                        number = list(number[number.index('-') + 1:]) #? Invoice numbers only include digits and 'I-' is skipped.
                        number = "".join(number)
                        invoice.invoice_number = number
                    else: #? '\n' not found it means cells are not merged and we can just pick the value from the cell directly
                        number = str(invoice_sheet['F2'].value)
                        number = list(number[number.index('-') + 1:])
                        number = "".join(number)
                        invoice.invoice_number = number
                else: # ? This means data cell is shifted towards right and now we have to pick the data from Cell G.
                    if '\n' in str(invoice_sheet['G1'].value):
                        number = str(invoice_sheet['G1'].value)
                        number = list(number[number.index('-') + 1:])
                        number = "".join(number)
                        invoice.invoice_number = number
                    else:
                        number = str(invoice_sheet['G2'].value)
                        number = list(number[number.index('-') + 1:])
                        number = "".join(number)
                        invoice.invoice_number = number
                
                #? Assigning NTN
                if '\n' in str(invoice_sheet['B2'].value):
                    NTN = str(invoice_sheet['B2'].value)
                    NTN = list(NTN[NTN.index('\n') + 1:])
                    NTN = "".join(NTN)
                    if len(NTN) == 8 or len(NTN) == 7 or len(NTN) == 9:
                        invoice.NTN = NTN
                else:
                    NTN = str(invoice_sheet['B2'].value)
                    if len(NTN) == 8 or len(NTN) == 7 or len(NTN) == 9:
                        invoice.NTN = NTN                                   
        elif 'Product Code' in cell_value_A1: #! --> Quantity's Sheet & Sales Tax Sheet
            SalesTax = 0
            last_sheet_no = len(invoice_sheet['B'])
            if checkInt(invoice_sheet[f'A{last_sheet_no}'].value) == True:
                for product_i in range(2, len(invoice_sheet['B']) + 1): #? This will iterate it through the product names.
                    product_name = invoice_sheet[f'B{product_i}'].value
                    SalesTax += invoice_sheet[f'K{product_i}'].value
                    
                    if "1" in product_name and "X 5" in product_name or "1" in product_name and "x 5" in product_name or "1" in product_name and "x5" in product_name or "1" in product_name and "X5" in product_name:
                        invoice.total_quantity += invoice_sheet[f'E{product_i}'].value #? if x 5 found, then Add Ctn quantity
                    else: 
                        invoice.total_quantity += invoice_sheet[f'F{product_i}'].value #? Add pcs quantity
            elif checkInt(invoice_sheet[f'A{last_sheet_no}'].value) == False:
                for product_i in range(2, len(invoice_sheet['B'])): #? This will iterate it through the product names.
                    product_name = invoice_sheet[f'B{product_i}'].value
                    SalesTax += invoice_sheet[f'K{product_i}'].value
                    
                    if "1" in product_name and "X 5" in product_name or "1" in product_name and "x 5" in product_name or "1" in product_name and "x5" in product_name or "1" in product_name and "X5" in product_name:
                        invoice.total_quantity += invoice_sheet[f'E{product_i}'].value #? if x 5 found, then Add Ctn quantity
                    else: 
                        invoice.total_quantity += invoice_sheet[f'F{product_i}'].value #? Add pcs quantity

            invoice.ValueAfterTax = SalesTax                 
        else: #! --> Sales meme invoice & shop information Sheet --> Useless & Time waste
            continue
        
        #* Important
        sheet_divider = (sheet_divider % 2) + 1  #? (1 % 3) + 1 = 2, (2 % 3) + 1 = 3, (3 % 3) + 1 = 1 --> We are using '% 3' because we are reading data from three sheets.
        if sheet_divider == 1: #? (2 != 1), (3 != 1), (1 == 1)
            Invoice_objects.append(invoice)
            invoice = invoicee()
    
    loader.stop()
    
    #? Exporting the object's values into the excel file...
    # * File to export
    loader = Loader("Looking into the sample file...", "Loaded Sample File!", 0.05).start()
    export_sheet = load_workbook('excel_files/sample.xlsx') #? File to Save later on
    to_export_sheet = export_sheet.active
    loader.stop()
    
    loader = Loader("Information stored in objects is being fetched into the excel file...", "Transfer Successful!", 0.05).start()
    rows = 2
    
    for i in range(0, len(Invoice_objects)):
        if len(Invoice_objects[i].NTN) > 1:
                to_export_sheet[f'D{rows}'].value = Invoice_objects[i].buyer_name #? Col D
                to_export_sheet[f'B{rows}'].value = Invoice_objects[i].NTN #? Col B
                to_export_sheet[f'H{rows}'].value = int(Invoice_objects[i].invoice_number) #? Col H
                to_export_sheet[f'I{rows}'].value = Invoice_objects[i].invoice_date #? Col I
                to_export_sheet[f'M{rows}'].value = Invoice_objects[i].total_quantity #? Col M
                to_export_sheet[f'O{rows}'].value = int(Invoice_objects[i].ValueAfterTax) #? Col O
        else:
            if Invoice_objects[i].buyer_cnic == '0' or Invoice_objects[i].buyer_cnic == 'None': #? If buyer cnic is empty it will search for it in the fbr.xlsx
                Invoice_objects[i].buyer_cnic = invoicee.ctrl_f(cur, Invoice_objects[i].buyer_name, Invoice_objects[i].buyer_cnic, fbr_shop_names, fbr_shop_cnics) #? Searches for cnic in fbr.xlsx
                if Invoice_objects[i].buyer_cnic == '0' or Invoice_objects[i].buyer_cnic == 'None' or Invoice_objects[i].buyer_cnic == '': #? If still not found we don't want it.
                    continue    
            #? If buyer_name found in fbr.xlsx and buyer_cnic is empty then it takes the cnic from the fbr file.
            to_export_sheet[f'D{rows}'].value = Invoice_objects[i].buyer_name #? Col D
            to_export_sheet[f'C{rows}'].value = Invoice_objects[i].buyer_cnic #? Col C
            to_export_sheet[f'H{rows}'].value = int(Invoice_objects[i].invoice_number) #? Col H
            to_export_sheet[f'I{rows}'].value = Invoice_objects[i].invoice_date #? Col I
            to_export_sheet[f'M{rows}'].value = Invoice_objects[i].total_quantity #? Col M
            to_export_sheet[f'O{rows}'].value = int(Invoice_objects[i].ValueAfterTax) #? Col O
        
        
    
        rows += 1
        
    #? Saves it. 
    loader.stop()
    conn.commit() #? Updates the fbr file as well
    conn.close() #! To close the connection
    export_sheet.save("Generated_FBR.xlsx")
    loader = Loader("Saving & Committing everything...", "File generated successfully & FBR database updated as well!", 0.05).start()
    loader.stop()

if __name__ == "__main__":
    start = default_timer() #? Starts a timer
    program()
    stop = default_timer() #? Stops a timer.
    
    total = stop - start #? Total time program took.
    total = round(total) #? Value is returned in highly precised float method thats why we want to round it.
    
    print(f"\nYour program took total of {total} seconds to complete...")
    input("Press any key to continue...") #? So the user can know if his program is completed or not.