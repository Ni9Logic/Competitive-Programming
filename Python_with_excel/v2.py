from timeit import default_timer
from openpyxl import load_workbook
from contextlib import closing
import openpyxl
import pymysql
import datetime

class invoicee:
    def __init__(self):
        self.buyer_name = ''
        self.buyer_cnic = ''
        self.invoice_date = ''
        self.invoice_number = 0
        self.total_quantity = 0
        self.ValueAfterTax = 0
        self.NTN = ''   
        
    def correct_buyer_names(self):
        replacements = {'G/S': 'GENERAL STORE', 'g/s': ' GENERAL STORE', 'S/S': 'SUPER STORE', 'C/C': 'CASH AND CARRY',
                        'K/S': 'KARYANA STORE', 'G.STORE': 'GENERAL STORE', 'C&C': 'CASH AND CARRY', ' GS': 'GENERAL STORE',
                        'GS': ' GENERAL STORE', 'gs': ' GENERAL STORE', 'Cash & Carry': 'CASH AND CARRY',
                        'Cash & carry': 'CASH AND CARRY', 'cash & carry': 'CASH AND CARRY', 'W/S': 'WHOLE SALE', 'w/s': 'WHOLE SALE',
                        'TRD.': 'TRADER', 'trd.': 'TRADER', 'TRD': 'TRADER', 'trd': 'TRADER'}
        
        for key, value in replacements.items():
            self.buyer_name = self.buyer_name.replace(key, value)
        return self.buyer_name
    def correct_dates(self):
        date_format = '%d-%m-%Y'
        if '00:00:00' in self.invoice_date:
            self.invoice_date = self.invoice_date.replace('00:00:00', '')

        self.invoice_date = datetime.strptime(
            self.invoice_date, date_format).strftime('%d-%b-%Y')

        if self.invoice_date.count('0') > 3:
            extra_zeroes = self.invoice_date[1:5]
            self.invoice_date = self.invoice_date.replace(extra_zeroes, '')
        return self.invoice_date
    
    def ctrl_f(cur, shop_name : str, shop_cnic : str, fbr_shop_names : list, fbr_shop_cnics : list) -> str:
        query = "INSERT INTO fbr (Buyer_CNIC, Buyer_Name) VALUES (%s, %s)"
    
        for fbr_shop_name, fbr_shop_cnic in zip(fbr_shop_names, fbr_shop_cnics):
            if fbr_shop_name is None or fbr_shop_cnic is None:
                continue
            if shop_name == fbr_shop_name and shop_cnic == '0' or shop_cnic == 'None':
                fbr_shop_cnic = fbr_shop_cnic.replace('-', '')
                shop_cnic = fbr_shop_cnic
            elif shop_name not in fbr_shop_names and shop_cnic != '0': #? This is just to update the FBR database with time to time.
                cur.execute(query, (shop_cnic, shop_name))
                break
            #? Import data and update from the database.
            
        return shop_cnic          

def checkInt(str):
    try:
        int(str)
        return True
    except ValueError:
        return False
def program():

    # Use the 'with' statement to automatically close the workbook after use
    converted = openpyxl.load_workbook('excel_files/invoice.xlsx')
    invoice_sheet = converted.active

    invoice = invoicee()
        # Add the invoice object to the list of objects
    Invoice_objects = [invoice]
    
    # * 1 --> Length of all the worksheets we got after converting the pdf into excel.
    for i, table_name in enumerate(converted.sheetnames, start=1):
        # ? To get all the data in invoice_sheet's object from current table.
        invoice_sheet = converted[table_name]
        # * This makes it further identifiable on which sheet we are currently.
        cell_value_A1 = invoice_sheet['A1'].value
        
        if cell_value_A1 is None or cell_value_A1 == 'Total No. Of Items 20':
            continue
        if invoice_sheet['C1'].value is None: #? CASH MEMO / SALE INVOICE WALI SHEET FFS
            buyer_name = invoice_sheet['C3'].value
            if '\n' in buyer_name:
                buyer_name = "".join(buyer_name[buyer_name.index('-') + 1:buyer_name.index('\n')]) #? On new line is address we don't want that and there's a dash as well
            else:                                                                                  
                buyer_name = "".join(buyer_name[buyer_name.index('-') + 1::])   #? We don't want the dash in the file.

            if buyer_name := list(invoice_sheet['C1'].value):
                invoice.buyer_name = buyer_name.upper()
                invoice.buyer_name = invoice.correct_buyer_names()
        else:
            buyer_name = invoice_sheet['C1'].value
            if '\n' in buyer_name:
                buyer_name = buyer_name.split('\n')[0]  # Get the first line of the name
            # Get the second part after the dash and remove whitespace
            buyer_name = buyer_name.split('-')[1].strip()
            buyer_name = buyer_name.upper()
            invoice.buyer_name = invoice.correct_buyer_names(buyer_name)

            
        # ? Assigning CNIC to object
        buyer_cnic_cell = invoice_sheet['C6'] if cell_value_A1 == "CASH MEMO / SALE INVOICE" else invoice_sheet['C4']
        buyer_cnic = str(buyer_cnic_cell.value).replace('_', '').replace('-', '') if buyer_cnic_cell.value else '0'
        invoice.buyer_cnic = buyer_cnic
                 
        # ? Assigning Dates to object
        if "Date" in str(invoice_sheet['R3'].value):
            date_cell = invoice_sheet['T3']
        elif "Date" in str(invoice_sheet['S1'].value):
            date_cell = invoice_sheet['U1']
        elif invoice_sheet['Q1'].value is not None and "Date" in str(invoice_sheet['Q1'].value):
            date_cell = invoice_sheet['R1']
        else:
            date_cell = None
        if date_cell is not None:
            date = str(date_cell.value)
        if '\n' in date:
            date = date.split('\n')[1]  # split to get only the date
            date = date.replace('/', '-')
            year = date[:4]
            day = date[8:]
            date = f"{year}-{date[5:7]}-{day}"
            invoice.invoice_date = date
            invoice.invoice_date = invoice.correct_dates()
            
        # ? Assigning Invoice Numbers
        # Define a list of possible cell locations for invoice number and date
        locations = [
            ('R3', 'T3'),
            ('R3', 'T4'),
            ('Q1', 'U1'),
            ('Q1', 'U2'),
            ('S1', 'S1'),
            ('S1', 'S2')
        ]
        # Loop through the locations and check if invoice number and date are present
        for loc in locations:
            if "Date" in str(invoice_sheet[loc[0]].value):
                if '\n' in str(invoice_sheet[loc[1]].value):
                    number = str(invoice_sheet[loc[1]].value)
                    number = list(number[number.index('-') + 1:])
                    number = "".join(number)
                    invoice.invoice_number = number
                else:
                    number = str(invoice_sheet[loc[1]+1].value)
                    number = list(number[number.index('-') + 1:])
                    number = "".join(number)
                    invoice.invoice_number = number
                break  # Exit the loop once invoice number is found
        # If invoice number is still not found, set it to None
        if invoice.invoice_number is None:
            invoice.invoice_number = "N/A"
            
        #? Assigning NTN
        ntn_value = str(invoice_sheet['C2'].value)
        if ntn_value:
            ntn_list = ntn_value.split('\n')
            if len(ntn_list) > 1:
                ntn = ntn_list[1]
                if len(ntn) in [7, 8, 9]:
                    invoice.NTN = ntn
            elif len(ntn_value) in [7, 8, 9]:
                invoice.NTN = ntn_value

        #? Calculating TotalQuantity and SalesTax  
        invoice.total_quantity += sum([invoice_sheet[f'G{start}'].value if "1" in invoice_sheet[f'B{start}'].value and ("X 5" in invoice_sheet[f'B{start}'].value.lower(
        ) or "x5" in invoice_sheet[f'B{start}'].value.lower()) else invoice_sheet[f'I{start}'].value for start in range(14, len(invoice_sheet['A']) + 1) if checkInt(invoice_sheet[f'A{start}'].value)])
        
        product_name = invoice_sheet[f'B{start}'].value.lower()
        if "1" in product_name and ("x 5" in product_name or "x5" in product_name or "x5" in product_name):
            # ? if x 5 found, then Add Ctn quantity
            invoice.total_quantity += invoice_sheet[f'G{start}'].value
        elif "1" in product_name:
            # ? Add pcs quantity
            invoice.total_quantity += invoice_sheet[f'I{start}'].value   
        if invoice_sheet['A1'].value == 'CASH MEMO / SALE INVOICE':
            start = 14
            SalesTax = sum([invoice_sheet[f'P{start}'].value for start in range(start, len(
                invoice_sheet['A']) + 1) if checkInt(invoice_sheet[f'A{start}'].value)])
        else:
            if invoice_sheet['P11'].value == "Further Tax" and invoice_sheet['P11'].value is not None:
                start = 12
                SalesTax = sum([invoice_sheet[f'N{start}'].value for start in range(start, len(
                    invoice_sheet['A']) + 1) if checkInt(invoice_sheet[f'A{start}'].value)])
            else:
                start = 12
                SalesTax = sum([invoice_sheet[f'P{start}'].value for start in range(start, len(
                    invoice_sheet['A']) + 1) if checkInt(invoice_sheet[f'A{start}'].value)])
        invoice.ValueAfterTax = SalesTax
        
        Invoice_objects.append(invoice)
        invoice = invoicee()
                                       
    
    #? Exporting the object's values into the excel file...
    # * File to export
    export_sheet = load_workbook('excel_files/sample.xlsx') #? File to Save later on
    to_export_sheet = export_sheet.active
    
    rows = 2
    
    for i in range(0, len(Invoice_objects)):
        if len(Invoice_objects[i].NTN) > 1:
                to_export_sheet[f'D{rows}'].value = Invoice_objects[i].buyer_name #? Col D
                to_export_sheet[f'B{rows}'].value = Invoice_objects[i].NTN #? Col B
                to_export_sheet[f'H{rows}'].value = int(Invoice_objects[i].invoice_number) #? Col H
                to_export_sheet[f'I{rows}'].value = Invoice_objects[i].invoice_date #? Col I
                to_export_sheet[f'M{rows}'].value = Invoice_objects[i].total_quantity #? Col M
                to_export_sheet[f'O{rows}'].value = int(Invoice_objects[i].ValueAfterTax) #? Col O
        else:
            if Invoice_objects[i].buyer_cnic == '0' or Invoice_objects[i].buyer_cnic == 'None': #? If buyer cnic is empty it will search for it in the fbr.xlsx
                Invoice_objects[i].buyer_cnic = invoicee.ctrl_f(cur, Invoice_objects[i].buyer_name, Invoice_objects[i].buyer_cnic, fbr_shop_names, fbr_shop_cnics) #? Searches for cnic in fbr.xlsx
                if Invoice_objects[i].buyer_cnic == '0' or Invoice_objects[i].buyer_cnic == 'None' or Invoice_objects[i].buyer_cnic == '': #? If still not found we don't want it.
                    continue    
            #? If buyer_name found in fbr.xlsx and buyer_cnic is empty then it takes the cnic from the fbr file.
            to_export_sheet[f'D{rows}'].value = Invoice_objects[i].buyer_name #? Col D
            to_export_sheet[f'C{rows}'].value = Invoice_objects[i].buyer_cnic #? Col C
            to_export_sheet[f'H{rows}'].value = int(Invoice_objects[i].invoice_number) #? Col H
            to_export_sheet[f'I{rows}'].value = Invoice_objects[i].invoice_date #? Col I
            to_export_sheet[f'M{rows}'].value = Invoice_objects[i].total_quantity #? Col M
            to_export_sheet[f'O{rows}'].value = int(Invoice_objects[i].ValueAfterTax) #? Col O
        
        
    
        rows += 1
        
    export_sheet.save("Generated_FBR.xlsx")

if __name__ == "__main__":
    start = default_timer() #? Starts a timer
    program()
    stop = default_timer() #? Stops a timer.
    
    total = stop - start #? Total time program took.
    total = round(total) #? Value is returned in highly precised float method thats why we want to round it.
    
    print(f"\nYour program took total of {total} seconds to complete...")
    # input("Press any key to continue...") #? So the user can know if his program is completed or not.